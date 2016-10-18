# -*- coding: utf-8 -*-
"""
Created on Tue Oct 11 22:48:47 2016

@author: Julio
"""
from openpyxl import load_workbook, worksheet

print("Bienvenido Carga de Datos")
try:
    #Base de datos1 - proyecto No. 1
    #mini_base
    wb = load_workbook('Base de datos1 - proyecto No. 1.xlsx', use_iterators=True, data_only=True)
    ws = wb.get_sheet_by_name('Hoja1')
except:
    print ("error al cargar archivo")
    
#Guardamos la base de datos en dos listas que comparten indice
ingresos=[]
fechas=[]

#Definimos zona de datos
filas = range(ws.min_row+1, ws.max_row+1)
columnas = range(ws.min_column, ws.max_column+1)

print ("Cargando Base de Datos. Por favor espere... ")
for fila in filas:
    try:
        #el ingreso se encuentra en segunda columna
        ingresos.append(ws.cell(row=fila, column=columnas[2]).value)
        #la fecha se encuentra en primera columna
        fechas.append(ws.cell(row=fila, column=columnas[1]).value.date())
    except:
        print (fila,"no es fecha")

print ("Escribiendo Base de Datos")
  
txt=open('Base_Datos.txt','w')
for i in range(0,len(fechas)):
    txt.write(((str)(fechas[i]))+','+((str)(ingresos[i]))+'\n')
txt.close()