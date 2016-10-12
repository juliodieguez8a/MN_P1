# -*- coding: utf-8 -*-
"""
Created on Sun Oct  9 22:01:12 2016

@author: Julio
"""

from pylab import *
from openpyxl import load_workbook, worksheet
from datetime import datetime


wb = load_workbook('Base de datos1 - proyecto No. 1.xlsx', use_iterators=True, data_only=True)
ws = wb.get_sheet_by_name('Hoja1')

op=input("Desea cargar base de datos (si/no): ")

if(op=="si"):

    #Guardamos la base de datos en dos listas que comparten indice
    ingresos=[]
    fechas=[]
    
    #Definimos zona de datos
    filas = range(ws.min_row+1, ws.max_row+1)
    columnas = range(ws.min_column, ws.max_column+1)
    
    print ("Cargando Base de Datos. Por favor espere... ")
    for fila in filas:
        #el ingreso se encuentra en segunda columna
        ingresos.append(ws.cell(row=fila, column=columnas[2]).value)
        #la fecha se encuentra en primera columna
        fechas.append(ws.cell(row=fila, column=columnas[1]).value)


print ("Se mostraran los ingresos en el intervalo de tiempo seleccionado\n"+
    "la fecha debe encontrarse entre 01/01/2012 y 31/07/2015")
#Ingreso de fechas limite
f_min=input("Ingrese fecha de inicio dd/mm/aaaa: ")
f_min=datetime.strptime(f_min,"%d/%m/%Y")
f_max=input("Ingrese fecha de final  dd/mm/aaaa: ")
f_max=datetime.strptime(f_max,"%d/%m/%Y")

#determinar el indice de la fechas maxima y minima
i_min=fechas.index(f_min)
i_max=fechas.index(f_max)

#la grafica tiene como eje x el rango dado por los indices
x=range (i_min, i_max+1)

#la grafica tiene en y el ingreso
y=[]
y2=[]

#inicializar variables
ingreso=0
ingreso_max=0.0

for i in x:
    ingreso_actual=ingresos[i]
    
    #tomamos el valor maximo para la grafica
    if ingreso_actual != None:
        ingreso_max=max(ingreso_max,ingreso_actual)
    y2.append(ingreso_actual)    
    
#==============================================================================
#     #COLAPSA CUANDO SON MUCHOS DATOS
#     #ingreso es lo que lleva mas el nuevo ingreso
#     ingreso+=ingreso_actual
#     y.append(ingreso)
#==============================================================================
    


#==============================================================================
# #GRAFICAR
# plot(x,y)
# #Definir caracteristicas de grafica
# axis([i_min,i_max+1,0,ingreso])
# grid()
# title("ingresos acumulados entre\n"+str(f_min)+" y "+str(f_max))
# xlabel("x")
# ylabel("y")
# 
# show()
#==============================================================================

#CREAR FIGURA
figure(1)
#GRAFICAR
plot(x,y2)
#Definir caracteristicas de grafica
axis([i_min,i_max+1,0,ingreso_max])
grid()
title("ingresos  entre\n"+str(f_min)+" y "+str(f_max))
xlabel("x")
ylabel("y")

savefig('foo.png')

show()